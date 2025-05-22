"""
Excel入力欄特定とJSON化
"""

import os
import json
import base64
import logging
from pathlib import Path
from typing import Dict, List, Any, TypedDict, Annotated, Literal, Optional

# LangChain関連のインポート
from langchain_core.messages import HumanMessage, SystemMessage
from langchain_core.prompts import ChatPromptTemplate
from langchain_openai import ChatOpenAI
from langgraph.graph import StateGraph, END
from pydantic import BaseModel, Field

# Excel操作関連のインポート
import openpyxl
from openpyxl.styles import PatternFill
import subprocess

# 環境変数の読み込み
from dotenv import load_dotenv
load_dotenv()

logger = logging.getLogger(__name__)

# Pydanticモデル: 入力欄情報
class ExcelField(BaseModel):
    """Excelの入力欄情報を表すモデル"""
    cell_id: str = Field(..., description="セル番号（例: A1, B2）")
    description: str = Field(..., description="そのセルに記入すべき内容の説明")

class ExcelFormFields(BaseModel):
    """Excelフォームの入力欄情報のコレクション"""
    fields: List[ExcelField] = Field(..., description="検出された入力欄のリスト")

# Pydanticモデル: 検証結果
class ValidationResult(BaseModel):
    """入力欄の検証結果を表すモデル"""
    status: Literal["OK", "修正が必要"] = Field(..., description="検証結果のステータス")
    issues: Optional[List[str]] = Field(None, description="問題点のリスト（ステータスが「修正が必要」の場合）")
    suggestions: Optional[List[str]] = Field(None, description="修正提案のリスト（ステータスが「修正が必要」の場合）")

# 状態の型定義
class ExcelFormState(TypedDict):
    excel_file: str
    output_dir: Optional[Path]
    max_iterations: int
    current_iteration: int
    extracted_text_file: str
    original_excel_capture: str
    estimated_fields: Dict[str, str] 
    structured_fields: ExcelFormFields
    highlighted_excel: str
    highlighted_captures: List[str]
    validation_result: str 
    structured_validation: ValidationResult
    validation_status: Literal["OK", "修正が必要", "エラー"]
    final_json: str
    status: Literal["進行中", "完了", "エラー"]
    error_message: str

# 1. Excelデータのテキスト化と画像キャプチャ
def extract_excel_data_and_capture(state: ExcelFormState) -> ExcelFormState:
    """
    Excelファイルからテキストデータを抽出し、画像キャプチャを取得する
    """
    logger.info(f"Excelテキスト抽出と画像キャプチャ開始: {state['excel_file']}")
    
    try:
        # 実際の保存先ベースディレクトリを決定
        user_defined_output_dir = state.get("output_dir")
        if user_defined_output_dir and str(user_defined_output_dir).strip():
            base_save_path = user_defined_output_dir
        else:
            base_save_path = Path(state["excel_file"]).parent
        
        final_output_dir = base_save_path / "format_data"
        final_output_dir.mkdir(exist_ok=True, parents=True)
        
        # キャプチャ用ディレクトリの作成 (final_output_dir の下に captures を作成)
        captures_dir = final_output_dir / "captures"
        captures_dir.mkdir(exist_ok=True, parents=True)
        
        # Excelファイルを開く
        workbook = openpyxl.load_workbook(state["excel_file"])
        
        # 抽出結果を格納するテキスト
        extracted_text = ""
        
        # 各シートの処理
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            
            # シート名の追加
            extracted_text += f"## シート名: {sheet_name}\n"
            
            # 結合セル情報の抽出
            merged_cells = []
            for merged_cell_range in sheet.merged_cells.ranges:
                merged_cells.append(str(merged_cell_range))
            
            if merged_cells:
                extracted_text += "### 結合セル情報:\n"
                for cell_range in merged_cells:
                    extracted_text += f"- {cell_range}\n"
            
            # セルデータの抽出
            extracted_text += "### セルデータ:\n"
            extracted_text += "| セル | 値 | 書式 |\n"
            extracted_text += "|-----|----|--------|\n"
            
            for row in sheet.iter_rows():
                for cell in row:
                    # セルが空でない場合のみ処理
                    if cell.value is not None:
                        cell_addr = f"{cell.column_letter}{cell.row}"
                        cell_value = str(cell.value)
                        
                        # 書式情報の取得
                        format_info = []
                        if cell.font.bold:
                            format_info.append("太字")
                        if cell.fill.fill_type == "solid":
                            fill_color = cell.fill.start_color.index
                            if fill_color != "00000000":  # デフォルト色でない場合
                                format_info.append(f"背景色:{fill_color}")
                        
                        format_str = ", ".join(format_info) if format_info else "-"
                        
                        # テーブルに行を追加
                        extracted_text += f"| {cell_addr} | {cell_value} | {format_str} |\n"
        
        # 抽出結果をファイルに保存
        extracted_text_file = final_output_dir / "extracted_excel_text.md"
        with open(extracted_text_file, "w", encoding="utf-8") as f:
            f.write(extracted_text)
        
        logger.info(f"Excelテキスト抽出完了: {extracted_text_file}")
        
        # 元のExcelファイルのキャプチャを取得
        # 一時ファイルにコピー (final_output_dir の下に保存)
        temp_excel = final_output_dir / "original_excel.xlsx"
        with open(state["excel_file"], "rb") as src, open(temp_excel, "wb") as dst:
            dst.write(src.read())
        
        # LibreOfficeを使用してPNGに変換
        command = f"soffice --headless --convert-to png {temp_excel} --outdir {captures_dir}"
        logger.info(f"実行コマンド: {command}")
        subprocess.run(command, shell=True, check=True)
        
        # 生成されたPNGファイルのパスを取得
        original_capture = captures_dir / "original_excel.png"
        if not original_capture.exists():
            # ファイル名が変更されている可能性があるため、キャプチャディレクトリ内のPNGファイルを探す
            png_files = list(captures_dir.glob("*.png"))
            if png_files:
                original_capture = png_files[0]
                # 標準的な名前にリネーム
                new_path = captures_dir / "original_excel.png"
                os.rename(original_capture, new_path)
                original_capture = new_path
        
        logger.info(f"元Excelのキャプチャ完了: {original_capture}")
        
        # 状態の更新
        return {
            **state,
            "extracted_text_file": str(extracted_text_file),
            "original_excel_capture": str(original_capture),
            "status": "進行中"
        }
        
    except Exception as e:
        logger.error(f"Excelテキスト抽出と画像キャプチャエラー: {str(e)}")
        return {
            **state,
            "status": "エラー",
            "error_message": f"Excelテキスト抽出と画像キャプチャエラー: {str(e)}"
        }

# 2. マルチモーダルLLMによる入力欄の推定（structured_output使用）
def estimate_fields_with_multimodal_llm(state: ExcelFormState) -> ExcelFormState:
    """
    テキストデータと画像キャプチャを使用してマルチモーダルLLMで入力欄を推定する
    structured_outputを使用して確実に指定の形式で結果を受け取る
    """
    logger.info(f"マルチモーダルLLMによる入力欄推定開始 (v{state['current_iteration']})")
    
    try:
        # 抽出されたテキストを読み込む
        with open(state["extracted_text_file"], "r", encoding="utf-8") as f:
            extracted_text = f.read()
        
        # 画像をbase64エンコード
        with open(state["original_excel_capture"], "rb") as img_file:
            base64_image = base64.b64encode(img_file.read()).decode("utf-8")
        
        # マルチモーダルLLMクライアントの初期化（structured_output使用）
        llm = ChatOpenAI(
            model="gpt-4.1-mini",
            temperature=0
        ).with_structured_output(ExcelFormFields)
        
        # プロンプトの作成
        prompt = f"""
あなたはExcelフォームの入力欄を特定する専門家です。

以下はExcelファイルから抽出したテキスト情報と、そのExcelシートの画像です。
このExcelファイルは入力フォームであり、ユーザーが情報を入力するセルを特定してください。

テキスト情報:
{extracted_text}

入力欄の特徴：
- 空白セル
- ラベル（太字や背景色付きのセル）の隣や下にある空白セル
- 表形式の場合、ヘッダー行の下の空白セル
- 既に値が入力されているセルでも、それが例や初期値と思われる場合は入力欄として扱う

画像とテキスト情報の両方を参考にして、入力欄を特定してください。
"""
        
        # マルチモーダルLLMに問い合わせ
        response = llm.invoke([
            HumanMessage(content=[
                {"type": "text", "text": prompt},
                {
                    "type": "image_url",
                    "image_url": {
                        "url": f"data:image/png;base64,{base64_image}"
                    }
                }
            ])
        ])
        
        # 構造化された応答を取得
        structured_fields = response
        
        # 従来の形式（Dict[str, str]）に変換（互換性のため）
        estimated_fields = {}
        for field in structured_fields.fields:
            estimated_fields[field.cell_id] = field.description
        
        # 実際の保存先ベースディレクトリを決定
        user_defined_output_dir = state.get("output_dir")
        if user_defined_output_dir and str(user_defined_output_dir).strip():
            base_save_path = user_defined_output_dir
        else:
            base_save_path = Path(state["excel_file"]).parent
        
        final_output_dir = base_save_path / "format_data"
        final_output_dir.mkdir(exist_ok=True, parents=True)
        
        # 構造化された形式を保存
        structured_fields_file = final_output_dir / f"structured_fields_v{state['current_iteration']}.json"
        with open(structured_fields_file, "w", encoding="utf-8") as f:
            f.write(structured_fields.model_dump_json(indent=2))
        
        # 従来の形式も保存（互換性のため）
        estimated_fields_file = final_output_dir / f"estimated_fields_v{state['current_iteration']}.json"
        with open(estimated_fields_file, "w", encoding="utf-8") as f:
            json.dump(estimated_fields, f, ensure_ascii=False, indent=2)
        
        logger.info(f"マルチモーダルLLMによる入力欄推定完了: {structured_fields_file}")
        
        # 状態の更新
        return {
            **state,
            "estimated_fields": estimated_fields,
            "structured_fields": structured_fields,
            "status": "進行中"
        }
        
    except Exception as e:
        logger.error(f"マルチモーダルLLMによる入力欄推定エラー: {str(e)}")
        return {
            **state,
            "status": "エラー",
            "error_message": f"マルチモーダルLLMによる入力欄推定エラー: {str(e)}"
        }

# 3. 入力欄のハイライト
def highlight_fields(state: ExcelFormState) -> ExcelFormState:
    """
    推定された入力欄をハイライトする
    """
    logger.info(f"入力欄のハイライト開始 (v{state['current_iteration']})")
    
    try:
        # 実際の保存先ベースディレクトリを決定
        user_defined_output_dir = state.get("output_dir")
        if user_defined_output_dir and str(user_defined_output_dir).strip():
            base_save_path = user_defined_output_dir
        else:
            base_save_path = Path(state["excel_file"]).parent
        
        final_output_dir = base_save_path / "format_data"
        final_output_dir.mkdir(exist_ok=True, parents=True)
        
        # 元のExcelファイルをコピー
        workbook = openpyxl.load_workbook(state["excel_file"])
        
        # 黄色のハイライト用フィル
        highlight_fill = PatternFill(
            start_color="FFFF00",
            end_color="FFFF00",
            fill_type="solid"
        )
        
        # 推定された入力欄をハイライト
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            
            for cell_addr in state["estimated_fields"].keys():
                try:
                    # セルアドレスが有効かチェック
                    if len(cell_addr) >= 2 and cell_addr[0].isalpha() and cell_addr[1:].isdigit():
                        cell = sheet[cell_addr]
                        cell.fill = highlight_fill
                except Exception as cell_error:
                    logger.warning(f"セル {cell_addr} のハイライト中にエラー: {str(cell_error)}")
        
        # ハイライト済みExcelを保存
        highlighted_excel = final_output_dir / f"highlighted_excel_v{state['current_iteration']}.xlsx"
        workbook.save(highlighted_excel)
        
        logger.info(f"入力欄のハイライト完了: {highlighted_excel}")
        
        # 状態の更新
        return {
            **state,
            "highlighted_excel": str(highlighted_excel),
            "status": "進行中"
        }
        
    except Exception as e:
        logger.error(f"入力欄のハイライトエラー: {str(e)}")
        return {
            **state,
            "status": "エラー",
            "error_message": f"入力欄のハイライトエラー: {str(e)}"
        }

# 4. ハイライト済みExcelのキャプチャ取得
def capture_highlighted_excel(state: ExcelFormState) -> ExcelFormState:
    """
    ハイライト済みExcelのキャプチャを取得する
    """
    logger.info(f"ハイライト済みExcelキャプチャ開始 (v{state['current_iteration']})")
    
    try:
        # 実際の保存先ベースディレクトリを決定
        user_defined_output_dir = state.get("output_dir")
        if user_defined_output_dir and str(user_defined_output_dir).strip():
            base_save_path = user_defined_output_dir
        else:
            base_save_path = Path(state["excel_file"]).parent
        
        final_output_dir = base_save_path / "format_data"
        final_output_dir.mkdir(exist_ok=True, parents=True)
        
        # captures ディレクトリは final_output_dir の下に作成
        captures_dir = final_output_dir / "captures"
        captures_dir.mkdir(exist_ok=True, parents=True)
        
        # LibreOfficeを使用してPNGに変換
        highlighted_excel = state["highlighted_excel"]
        command = f"soffice --headless --convert-to png {highlighted_excel} --outdir {captures_dir}"
        
        logger.info(f"実行コマンド: {command}")
        subprocess.run(command, shell=True, check=True)
        
        # 生成されたPNGファイルのパスを取得
        excel_filename = os.path.basename(highlighted_excel)
        excel_basename = os.path.splitext(excel_filename)[0]
        
        highlighted_captures = []
        for sheet_idx, sheet_name in enumerate(openpyxl.load_workbook(highlighted_excel).sheetnames, 1):
            capture_path = captures_dir / f"{excel_basename}_sheet{sheet_idx}.png"
            if not capture_path.exists():
                # ファイル名が変更されている可能性があるため、キャプチャディレクトリ内の最新のPNGファイルを探す
                png_files = list(captures_dir.glob("*.png"))
                if png_files:
                    # 最新のファイルを取得
                    latest_png = max(png_files, key=os.path.getctime)
                    # 標準的な名前にリネーム
                    os.rename(latest_png, capture_path)
            
            if capture_path.exists():
                highlighted_captures.append(str(capture_path))
        
        logger.info(f"ハイライト済みExcelキャプチャ完了: {highlighted_captures}")
        
        # 状態の更新
        return {
            **state,
            "highlighted_captures": highlighted_captures,
            "status": "進行中"
        }
        
    except Exception as e:
        logger.error(f"ハイライト済みExcelキャプチャ取得エラー: {str(e)}")
        return {
            **state,
            "status": "エラー",
            "error_message": f"ハイライト済みExcelキャプチャ取得エラー: {str(e)}"
        }

# 5. マルチモーダルLLMによる検証（structured_output使用）
def validate_with_multimodal_llm(state: ExcelFormState) -> ExcelFormState:
    """
    マルチモーダルLLMを使用してハイライト済み入力欄の検証を行う
    structured_outputを使用して確実に指定の形式で結果を受け取る
    """
    logger.info(f"マルチモーダルLLMによる検証開始 (v{state['current_iteration']})")
    
    try:
        # 実際の保存先ベースディレクトリを決定
        user_defined_output_dir = state.get("output_dir")
        if user_defined_output_dir and str(user_defined_output_dir).strip():
            base_save_path = user_defined_output_dir
        else:
            base_save_path = Path(state["excel_file"]).parent
        
        final_output_dir = base_save_path / "format_data"
        final_output_dir.mkdir(exist_ok=True, parents=True)
        
        # 推定された入力欄情報
        structured_fields = state["structured_fields"]
        
        # マルチモーダルLLMクライアントの初期化（structured_output使用）
        llm = ChatOpenAI(
            model="gpt-4.1-mini",
            temperature=0
        ).with_structured_output(ValidationResult)
        
        validation_results = []
        structured_validations = []
        
        for capture_path in state["highlighted_captures"]:
            # 画像をbase64エンコード
            with open(capture_path, "rb") as img_file:
                base64_image = base64.b64encode(img_file.read()).decode("utf-8")
            
            # プロンプトの作成
            prompt = f"""
以下は、Excelフォームの入力欄として推定されたセルをハイライト（黄色背景）した画像です。

このハイライトされた箇所について、以下の観点で評価を行ってください。
- 入力欄として適切なセルがハイライトされているか
- 入力すべきでない欄がハイライトされていないか

問題がなければステータスを「OK」としてください。
問題がある場合は、ステータスを「修正が必要」とし、具体的な問題点と修正案を説明してください。
"""
            
            # マルチモーダルLLMに問い合わせ
            response = llm.invoke([
                HumanMessage(content=[
                    {"type": "text", "text": prompt},
                    {
                        "type": "image_url",
                        "image_url": {
                            "url": f"data:image/png;base64,{base64_image}"
                        }
                    }
                ])
            ])
            
            # 構造化された検証結果を取得
            structured_validation = response
            structured_validations.append(structured_validation)
            
            # 従来の形式のテキスト応答も生成（互換性のため）
            validation_text = f"検証結果: {structured_validation.status}\n"
            if structured_validation.issues:
                validation_text += "問題点:\n" + "\n".join([f"- {issue}" for issue in structured_validation.issues]) + "\n"
            if structured_validation.suggestions:
                validation_text += "修正案:\n" + "\n".join([f"- {suggestion}" for suggestion in structured_validation.suggestions]) + "\n"
            
            validation_results.append(validation_text)
            
            # 検証結果をログに記録
            capture_filename = os.path.basename(capture_path)
            logger.info(f"検証結果 ({capture_filename}): {structured_validation.status}")
        
        # 検証結果をファイルに保存
        validation_result_file = final_output_dir / f"validation_result_v{state['current_iteration']}.txt"
        with open(validation_result_file, "w", encoding="utf-8") as f:
            f.write(f"シート {state['current_iteration']} の検証結果:\n")
            f.write("\n\n".join(validation_results))
        logger.info(f"検証結果ファイル保存: {validation_result_file}")
        
        # 構造化された検証結果を保存
        structured_validation_file = final_output_dir / f"structured_validation_v{state['current_iteration']}.json"
        with open(structured_validation_file, "w", encoding="utf-8") as f:
            # 複数の検証結果がある場合は最初のものを使用
            f.write(structured_validations[0].model_dump_json(indent=2))
        logger.info(f"構造化検証結果ファイル保存: {structured_validation_file}")
        
        # 検証結果の分析
        validation_status = "OK"
        if any(validation.status == "修正が必要" for validation in structured_validations):
            validation_status = "修正が必要"
        
        logger.info(f"検証完了: 結果={validation_status}")
        
        # 状態の更新
        return {
            **state,
            "validation_result": "\n\n".join(validation_results),
            "structured_validation": structured_validations[0],  # 複数ある場合は最初のものを使用
            "validation_status": validation_status,
            "status": "進行中"
        }
        
    except Exception as e:
        logger.error(f"検証エラー: {str(e)}")
        return {
            **state,
            "status": "エラー",
            "error_message": f"検証エラー: {str(e)}"
        }

# 6. 入力欄情報の修正（structured_output使用）
def correct_fields_with_multimodal_llm(state: ExcelFormState) -> ExcelFormState:
    """
    検証結果に基づいて入力欄情報を修正する
    structured_outputを使用して確実に指定の形式で結果を受け取る
    """
    logger.info(f"入力欄情報の修正開始 (v{state['current_iteration'] + 1})")
    
    try:
        # 実際の保存先ベースディレクトリを決定
        user_defined_output_dir = state.get("output_dir")
        if user_defined_output_dir and str(user_defined_output_dir).strip():
            base_save_path = user_defined_output_dir
        else:
            base_save_path = Path(state["excel_file"]).parent
        
        final_output_dir = base_save_path / "format_data"
        final_output_dir.mkdir(exist_ok=True, parents=True)
        
        # 現在の推定結果
        structured_fields = state["structured_fields"]
        
        # 検証結果
        structured_validation = state["structured_validation"]
        
        # ハイライトされたExcel画像
        with open(state["highlighted_captures"][0], "rb") as img_file:
            base64_image = base64.b64encode(img_file.read()).decode("utf-8")
        
        # マルチモーダルLLMクライアントの初期化（structured_output使用）
        llm = ChatOpenAI(
            model="gpt-4.1-mini",
            temperature=0
        ).with_structured_output(ExcelFormFields)
        
        # プロンプトの作成
        prompt = f"""
あなたはExcelフォームの入力欄を特定する専門家です。
以下のSTEPで作業をしてください。
STEP1:以下の情報をよく確認してください。
- 現在推定されているExcelフォームの入力欄情報
{structured_fields.model_dump_json(indent=2)}

- 添付の画像 ※推定された入力欄をハイライトしたExcelシートの画像

- これらに対するレビュー結果
{structured_validation.model_dump_json(indent=2)}

STEP2:検証結果と画像に基づいて、入力欄情報を修正してください。指摘がない箇所は変更しないで回答に含めてください。
"""
        
        # マルチモーダルLLMに問い合わせ
        response = llm.invoke([
            HumanMessage(content=[
                {"type": "text", "text": prompt},
                {
                    "type": "image_url",
                    "image_url": {
                        "url": f"data:image/png;base64,{base64_image}"
                    }
                }
            ])
        ])
        
        # 構造化された応答を取得
        corrected_structured_fields = response
        
        # 従来の形式（Dict[str, str]）に変換（互換性のため）
        corrected_fields = {}
        for field in corrected_structured_fields.fields:
            corrected_fields[field.cell_id] = field.description
        
        # 結果をファイルに保存
        next_iteration = state["current_iteration"] + 1
        
        # 構造化された形式を保存
        structured_fields_file = final_output_dir / f"structured_fields_v{next_iteration}.json"
        with open(structured_fields_file, "w", encoding="utf-8") as f:
            f.write(corrected_structured_fields.model_dump_json(indent=2))
        
        # 従来の形式も保存（互換性のため）
        corrected_fields_file = final_output_dir / f"estimated_fields_v{next_iteration}.json"
        with open(corrected_fields_file, "w", encoding="utf-8") as f:
            json.dump(corrected_fields, f, ensure_ascii=False, indent=2)
        
        logger.info(f"入力欄情報の修正完了: {structured_fields_file}")
        
        # 状態の更新
        return {
            **state,
            "estimated_fields": corrected_fields,
            "structured_fields": corrected_structured_fields,
            "current_iteration": next_iteration,
            "status": "進行中"
        }
        
    except Exception as e:
        logger.error(f"入力欄情報の修正エラー: {str(e)}")
        return {
            **state,
            "status": "エラー",
            "error_message": f"入力欄情報の修正エラー: {str(e)}"
        }

# 7. 最終結果の生成
def generate_final_json(state: ExcelFormState) -> ExcelFormState:
    """
    最終的な入力欄情報JSONを生成する
    """
    logger.info("最終結果の生成")
    
    try:
        # 実際の保存先ベースディレクトリを決定
        user_defined_output_dir = state.get("output_dir")
        if user_defined_output_dir and str(user_defined_output_dir).strip():
            base_save_path = user_defined_output_dir
        else:
            base_save_path = Path(state["excel_file"]).parent
        
        final_output_dir = base_save_path / "format_data"
        final_output_dir.mkdir(exist_ok=True, parents=True)
        
        # 最終的な入力欄情報
        final_fields = state["estimated_fields"]
        final_structured_fields = state["structured_fields"]
        
        # 結果をファイルに保存
        final_json_file = final_output_dir / "final_form_definition.json"
        with open(final_json_file, "w", encoding="utf-8") as f:
            json.dump(final_fields, f, ensure_ascii=False, indent=2)
        
        # 構造化された形式も保存
        final_structured_file = final_output_dir / "final_structured_form_definition.json"
        with open(final_structured_file, "w", encoding="utf-8") as f:
            f.write(final_structured_fields.model_dump_json(indent=2))
        
        logger.info(f"処理が完了しました。最終結果: {final_json_file}")
        
        # 状態の更新
        return {
            **state,
            "final_json": str(final_json_file),
            "status": "完了"
        }
        
    except Exception as e:
        logger.error(f"最終結果の生成エラー: {str(e)}")
        return {
            **state,
            "status": "エラー",
            "error_message": f"最終結果の生成エラー: {str(e)}"
        }

# ルーター関数: 検証結果に基づいて次のステップを決定
def router(state: ExcelFormState) -> str:
    """
    状態に基づいて次のステップを決定する
    """
    # エラーが発生した場合は終了
    if state["status"] == "エラー":
        return END
    
    # 処理が完了した場合は終了
    if state["status"] == "完了":
        return END
    
    # 検証結果に基づく分岐
    if state["validation_status"] == "OK":
        return "generate_final_json"
    
    # 最大反復回数に達した場合は最終結果を生成
    if state["current_iteration"] >= state["max_iterations"]:
        logger.warning(f"最大反復回数 ({state['max_iterations']}) に到達しました")
        return "generate_final_json"
    
    # 修正が必要な場合は修正ステップへ
    return "correct_fields_with_multimodal_llm"

# LangGraphワークフローの構築
def build_workflow() -> StateGraph:
    """
    Excel入力欄特定ワークフローを構築する
    """
    # グラフの作成
    workflow = StateGraph(ExcelFormState)
    
    # ノードの追加
    workflow.add_node("extract_excel_data_and_capture", extract_excel_data_and_capture)
    workflow.add_node("estimate_fields_with_multimodal_llm", estimate_fields_with_multimodal_llm)
    workflow.add_node("highlight_fields", highlight_fields)
    workflow.add_node("capture_highlighted_excel", capture_highlighted_excel)
    workflow.add_node("validate_with_multimodal_llm", validate_with_multimodal_llm)
    workflow.add_node("correct_fields_with_multimodal_llm", correct_fields_with_multimodal_llm)
    workflow.add_node("generate_final_json", generate_final_json)
    
    # エッジの追加（基本フロー）
    workflow.add_edge("extract_excel_data_and_capture", "estimate_fields_with_multimodal_llm")
    workflow.add_edge("estimate_fields_with_multimodal_llm", "highlight_fields")
    workflow.add_edge("highlight_fields", "capture_highlighted_excel")
    workflow.add_edge("capture_highlighted_excel", "validate_with_multimodal_llm")
    
    # 条件分岐
    workflow.add_conditional_edges(
        "validate_with_multimodal_llm",
        router,
        {
            "correct_fields_with_multimodal_llm": "correct_fields_with_multimodal_llm",
            "generate_final_json": "generate_final_json",
            END: END
        }
    )
    
    # 修正後のフロー
    workflow.add_edge("correct_fields_with_multimodal_llm", "highlight_fields")
    
    # 開始ノードの設定
    workflow.set_entry_point("extract_excel_data_and_capture")
    
    return workflow

